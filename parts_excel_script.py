from openpyxl import load_workbook
import sys
import xlwings as xw



input_accept = False
godzilla_parts_workbook ='parts_list.xlsx'

# Lets have some class and make a class, rather than having a nasty long script that does everything...all puns intended of course
class excelomator:
    # Initialize the workbook and excel sheet we're working on. The wbxl is to get evaluated formulae (the values) for cells
    # that contain formulas. OpenPyXL returns the formula as a string as the the value...?!
    def __init__(self,wbook):
        self.godzilla_workbook = load_workbook(wbook)
        self.godzilla_wbxl = xw.Book(wbook)
        self.ws = self.godzilla_workbook.active

    # This method will ask for the part, check if it's correct, and then add it to the spreadsheet. On the spreadsheet columns 2,3,4
    # correspond to Brand, Part, and Purchase Price
    def add_part(self):
        while True:
            # Get me inputs
            part_string = input("Add part using 'Brand, Part, Purchase $' format:  ")
            part_list = part_string.split(", ")

            # Make sure I didn't ... up
            print("You inputted: \n   Brand: " + part_list[0] + "\n   Part: " + part_list[1] + "\n   Purchase Price: " + part_list[2])

            keep_or_not = input("Do you want to keep this input? Answer y/n\n")

            if keep_or_not == "y":
                break
            elif keep_or_not != "y" and input_accept != "n":
                ("Input error, please type 'y' for yes and 'n' for no...Input part again\n")

        brand = part_list[0]
        part = part_list[1]
        price = part_list[2]

        max_row = 200
        added=False

        for row in range (6,max_row):
            for col in range(2, 5):
                cell = self.ws.cell(column=col, row=row)

                if cell.value is None:
                    added=True
                    if col == 2:
                        print("changing cell valu for brand")
                        cell.value = brand
                    elif col == 3:
                        cell.value = part
                    elif col == 4:
                        cell.value = int(price)
            if added:
                break
            
        
        # Now sum up price
        self.ws["F5"] = "=SUM(D6:D2000)"

    # Adios
    def peace_out(self):
        print("Nos vemos cuando gastas mas dinero ;)")

        self.godzilla_workbook.save("parts_list.xlsx")
        print("Workbook Saved")
        exit()

    # See the sad truth on how much I've spent
    def check_damage(self):
        print("\nPrepare yourself....: {}".format(self.godzilla_wbxl.sheets['Sheet1'].range('F5').value))

    # Dictionary to decide which method to run.
    def chooser(self,choice):
        switcher = {
                    1: self.add_part,
                    2: self.check_damage,
                    0: self.peace_out
                    }
        func = switcher.get(choice, lambda:'Invalid, choose from the listed choices')
        return func()
    

# Now we run the interface
interface = excelomator(godzilla_parts_workbook)
print("Welcome to the interface...")

while True:
    action_choice = int(input("What would you like to do?\n\n 1 - Add part \n 2 - See spending total \n 0 - Save and Exit\n\nInput: "))

    choice = interface.chooser(action_choice)